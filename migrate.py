#!/usr/bin/env python3
"""
FINANCE LAB — Migración COMPLETA de Excel a PostgreSQL
Fuente: Base_de_Gasto_y_Presupuesto_-_Bluecoins.xlsx (DATOS REALES)
Migra: gastos, cuentas, presupuestos, catálogos, registros borrados
"""

import openpyxl
import psycopg2
from datetime import datetime, date

DB = {
    "host": "localhost",
    "port": 5432,
    "dbname": "finance_lab",
    "user": "finadmin",
    "password": "f1nl4b_s3cur3",
}

# ── Ruta del XLSX REAL ──
import glob
XLSX_CANDIDATES = glob.glob("/root/.openclaw/media/inbound/Base_de_Gasto_y_Presupuesto_*.xlsx")
XLSX = XLSX_CANDIDATES[0] if XLSX_CANDIDATES else "/root/.openclaw/media/inbound/Base_de_Gasto_y_Presupuesto_-_Bluecoins---cbf92f67-e188-4018-a7dd-9e562034502c.xlsx"

SENTINEL_DATE = date(2026, 1, 1)

def connect():
    return psycopg2.connect(**DB)

def load_workbook():
    print(f"📂 Cargando: {XLSX}")
    return openpyxl.load_workbook(XLSX, data_only=True)

def safe_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        d = val.date()
        return None if d == SENTINEL_DATE else d
    if isinstance(val, date):
        return None if val == SENTINEL_DATE else val
    if isinstance(val, str):
        try:
            d = datetime.strptime(val, "%Y-%m-%d").date()
            return None if d == SENTINEL_DATE else d
        except:
            return None
    return None


# ═══════════════════════════════════════════════════════════════
# PASO 1: Limpiar datos de gastos existentes (migración anterior)
# ═══════════════════════════════════════════════════════════════

def clean_existing_data(cur):
    """Limpia solo los gastos y recordatorios viejos.
    Los catálogos y cuentas se actualizan (upsert), no se borran."""
    print("\n🧹 Limpiando datos existentes...")
    cur.execute("DELETE FROM gastos")
    cur.execute("DELETE FROM recordatorios")
    cur.execute("DELETE FROM gastos_borrados")
    print("   ✅ Gastos, recordatorios y papelera limpiados")


# ═══════════════════════════════════════════════════════════════
# PASO 2: Importar Cuentas desde Config_Cuentas
# ═══════════════════════════════════════════════════════════════

def seed_cuentas(cur, wb):
    ws = wb["Config_Cuentas"]
    headers = [c.value for c in ws[1]]
    
    cuentas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d["Nombre_Cuenta"]:
            cuentas.append(d)
    
    for c in cuentas:
        dia = int(c["Dia_Corte"]) if c["Dia_Corte"] and c["Dia_Corte"] > 0 else None
        cur.execute("""
            INSERT INTO cuentas (nombre, tipo, propietario, saldo_inicial, dia_corte, activo)
            VALUES (%s, %s, %s, %s, %s, true)
            ON CONFLICT (nombre) DO UPDATE SET
                tipo = EXCLUDED.tipo,
                propietario = EXCLUDED.propietario,
                saldo_inicial = EXCLUDED.saldo_inicial,
                dia_corte = EXCLUDED.dia_corte,
                activo = true,
                updated_at = now()
        """, (
            c["Nombre_Cuenta"],
            c["Tipo"] or "Activo",
            c["Propietario"] or "Josué",
            float(c["Saldo_Inicial"] or 0),
            dia,
        ))
    
    print(f"✅ {len(cuentas)} cuentas importadas/actualizadas")


# ═══════════════════════════════════════════════════════════════
# PASO 3: Importar Catálogos desde Config_Presupuesto
# ═══════════════════════════════════════════════════════════════

def seed_catalogos(cur, wb):
    """Del Config_Presupuesto extraemos:
    - TiposOperacion (ya existen: Gasto, Ingreso, Transferencia)
    - MacroCategorias
    - Categorias
    - categoria_arbol (tipo → macro → categoria)
    """
    ws = wb["Config_Presupuesto"]
    headers = [c.value for c in ws[1]]
    
    macro_set = set()
    cat_set = set()
    arbol = []  # (tipo_operacion, macro, categoria)
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if not d["Categoria"]:
            continue
        tipo = (d["TipoOperacion"] or "Gasto").strip()
        macro = (d["MacroCategoria"] or "General").strip()
        cat = d["Categoria"].strip()
        
        macro_set.add(macro)
        cat_set.add(cat)
        arbol.append((tipo, macro, cat))
    
    # También extraemos categorías de BaseRegistros por si hay alguna no listada en presupuesto
    ws2 = wb["BaseRegistros"]
    h2 = [c.value for c in ws2[1]]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        d = dict(zip(h2, row))
        if d["ID_Unico"] is None:
            continue
        if d["Categoria"]:
            cat_set.add(d["Categoria"].strip())
        if d["MacroCategoria"]:
            macro_set.add(d["MacroCategoria"].strip())
        if d["Tipo_Operacion"]:
            # Asegurar que existe en tipos_operacion
            cur.execute("INSERT INTO tipos_operacion (nombre) VALUES (%s) ON CONFLICT DO NOTHING",
                       (d["Tipo_Operacion"].strip(),))
    
    # Insertar macros
    for m in sorted(macro_set):
        cur.execute("INSERT INTO macro_categorias (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING", (m,))
    
    # Insertar categorías con iconos
    iconos = {
        "Alimentación y Despensa": "🛒",
        "Restaurantes y Cafeterías": "🍽️",
        "Vivienda y Servicios (Renta, Luz, Agua)": "🏠",
        "Suscripciones y Tecnología": "💻",
        "Transporte (Gasolina, Uber, Mto)": "🚗",
        "Mascotas (Gatos)": "🐱",
        "Citas Médicas y Medicamentos": "💊",
        "Cuidado Personal": "🧴",
        "Ahorro e Inversión": "💰",
        "Ropa y Calzado": "👕",
        "Muebles y Hogar": "🛋️",
        "Regalos (Familia, Amigos)": "🎁",
        "Vacaciones": "✈️",
        "Entretenimiento y Ocio": "🎬",
        "Gusguerías": "🍬",
        "Otros Gastos": "📦",
        "Gasolina y Transporte": "🚗",
        "Sueldo": "💵",
        "Otros Ingresos": "💸",
    }
    
    for cat in sorted(cat_set):
        cur.execute(
            "INSERT INTO categorias (nombre, icono) VALUES (%s, %s) ON CONFLICT (nombre) DO NOTHING",
            (cat, iconos.get(cat, "📌")),
        )
    
    print(f"✅ {len(macro_set)} macros, {len(cat_set)} categorías")
    
    # Reconstruir categoria_arbol desde cero con los datos reales
    cur.execute("DELETE FROM categoria_arbol")
    
    # Mapeo tipo_operacion nombre → id
    cur.execute("SELECT id, nombre FROM tipos_operacion")
    tipo_map = {row[1]: row[0] for row in cur.fetchall()}
    
    for tipo, macro, cat in sorted(arbol):
        tipo_id = tipo_map.get(tipo, tipo_map.get("Gasto", 1))
        cur.execute(
            "INSERT INTO categoria_arbol (tipo_operacion_id, macro, nombre) VALUES (%s, %s, %s) ON CONFLICT (tipo_operacion_id, macro, nombre) DO NOTHING",
            (tipo_id, macro, cat),
        )
    
    # También asegurar que toda categoría usada en BaseRegistros tenga entrada en arbol
    ws2 = wb["BaseRegistros"]
    h2 = [c.value for c in ws2[1]]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        d = dict(zip(h2, row))
        if d["ID_Unico"] is None:
            continue
        tipo = (d["Tipo_Operacion"] or "Gasto").strip()
        macro = (d["MacroCategoria"] or "General").strip()
        cat = (d["Categoria"] or "").strip()
        if not cat:
            continue
        tipo_id = tipo_map.get(tipo, tipo_map.get("Gasto", 1))
        cur.execute(
            "INSERT INTO categoria_arbol (tipo_operacion_id, macro, nombre) VALUES (%s, %s, %s) ON CONFLICT (tipo_operacion_id, macro, nombre) DO NOTHING",
            (tipo_id, macro, cat),
        )
    
    # Add unique constraint if it doesn't exist
    try:
        cur.execute("""
            DO $$ BEGIN
                IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'categoria_arbol_unique') THEN
                    ALTER TABLE categoria_arbol ADD CONSTRAINT categoria_arbol_unique 
                    UNIQUE (tipo_operacion_id, macro, nombre);
                END IF;
            END $$;
        """)
    except:
        pass
    
    count = cur.rowcount
    cur.execute("SELECT COUNT(*) FROM categoria_arbol")
    arbol_count = cur.fetchone()[0]
    print(f"✅ categoria_arbol reconstruido: {arbol_count} entradas")
    
    # Retornar los datasets para usar luego
    return {"macros": macro_set, "cats": cat_set}


# ═══════════════════════════════════════════════════════════════
# PASO 4: Importar Gastos desde BaseRegistros
# ═══════════════════════════════════════════════════════════════

def seed_gastos(cur, wb):
    ws = wb["BaseRegistros"]
    headers = [c.value for c in ws[1]]
    
    # Construir mapeos de lookup
    cur.execute("SELECT id, nombre FROM tipos_operacion")
    top_map = {row[1]: row[0] for row in cur.fetchall()}
    cur.execute("SELECT id, nombre FROM macro_categorias")
    mc_map = {row[1]: row[0] for row in cur.fetchall()}
    cur.execute("SELECT id, nombre FROM categorias")
    cat_map = {row[1]: row[0] for row in cur.fetchall()}
    cur.execute("SELECT id, nombre FROM tipos_gasto")
    tg_map = {row[1]: row[0] for row in cur.fetchall()}
    cur.execute("SELECT id, nombre FROM registradores")
    reg_map = {row[1]: row[0] for row in cur.fetchall()}
    cur.execute("SELECT id, nombre FROM cuentas")
    cuen_map = {row[1]: row[0] for row in cur.fetchall()}
    
    count = 0
    rec_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d["ID_Unico"] is None:
            continue
        
        # ── Determinar si es recordatorio ──
        es_rec = str(d["Es_Recordatorio"]).upper() if d["Es_Recordatorio"] is not None else "FALSE"
        
        # ── Tipos de Gasto que son "N/A" para transferencias ──
        tipo_gasto = (d["Tipo_Gasto"] or "").strip()
        if tipo_gasto == "N/A":
            tipo_gasto = "Compartido (50/50)"  # default fallback
        
        # ── Asegurar que existan en catálogos ──
        reg = (d["Registrador"] or "").strip()
        cat = (d["Categoria"] or "").strip()
        mac = (d["MacroCategoria"] or "").strip()
        top = (d["Tipo_Operacion"] or "Gasto").strip()
        co = (d["Cuenta_Origen"] or "").strip()
        cd = (d["Cuenta_Destino"] or "").strip()
        
        # find_or_create para registradores y tipos_gasto
        if reg:
            cur.execute("INSERT INTO registradores (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING", (reg,))
            if reg not in reg_map:
                cur.execute("SELECT id FROM registradores WHERE nombre = %s", (reg,))
                row_r = cur.fetchone()
                if row_r:
                    reg_map[reg] = row_r[0]
        
        if tipo_gasto and tipo_gasto not in tg_map:
            cur.execute("INSERT INTO tipos_gasto (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING", (tipo_gasto,))
            cur.execute("SELECT id FROM tipos_gasto WHERE nombre = %s", (tipo_gasto,))
            row_r = cur.fetchone()
            if row_r:
                tg_map[tipo_gasto] = row_r[0]
        
        # ── Valores ──
        reg_id = reg_map.get(reg) if reg else None
        cat_id = cat_map.get(cat) if cat else None
        mc_id = mc_map.get(mac) if mac else None
        top_id = top_map.get(top) if top else top_map.get("Gasto", 1)
        tg_id = tg_map.get(tipo_gasto) if tipo_gasto else None
        co_id = cuen_map.get(co) if co else None
        cd_id = cuen_map.get(cd) if cd else None
        
        fecha_compra = safe_date(d["Fecha_Compra"])
        fecha_reg = safe_date(d["Fecha_Registro"])
        monto_total = float(d["Monto_Total"] or 0)
        monto_parcial = float(d["Monto_Parcial"] or 0) if d["Monto_Parcial"] is not None else monto_total
        msi_status = safe_date(d["MSI_Status"])
        periodo_pago = safe_date(d["Periodo_Pago"])
        p_josue = float(d["Participacion_Josue"] or 0)
        p_abi = float(d["Participacion_Abi"] or 0)
        
        # ── Insertar ──
        if es_rec == "TRUE":
            cur.execute("""
                INSERT INTO recordatorios (
                    id_unico, fecha_programada, fecha_registro, registrador_id,
                    compra, categoria_id, macro_categoria_id,
                    monto_parcial, tipo_operacion_id,
                    cuenta_origen_id, cuenta_destino_id,
                    tipo_gasto_id, participacion_josue, participacion_abi
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id_unico) DO NOTHING
            """, (
                str(d["ID_Unico"]), fecha_compra, fecha_reg or date.today(), reg_id,
                d["Compra"] or "", cat_id, mc_id,
                monto_parcial, top_id, co_id, cd_id,
                tg_id, p_josue, p_abi,
            ))
            rec_count += 1
        else:
            cur.execute("""
                INSERT INTO gastos (
                    id_unico, fecha_compra, fecha_registro, registrador_id,
                    compra, categoria_id, macro_categoria_id,
                    monto_total, monto_parcial, msi_status,
                    tipo_gasto_id, periodo_pago,
                    participacion_josue, participacion_abi,
                    tipo_operacion_id, cuenta_origen_id, cuenta_destino_id
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON CONFLICT (id_unico) DO NOTHING
            """, (
                str(d["ID_Unico"]), fecha_compra, fecha_reg or date.today(), reg_id,
                d["Compra"] or "", cat_id, mc_id,
                monto_total, monto_parcial, msi_status,
                tg_id, periodo_pago,
                p_josue, p_abi,
                top_id, co_id, cd_id,
            ))
            count += 1
    
    print(f"✅ {count} gastos + {rec_count} recordatorios insertados")


# ═══════════════════════════════════════════════════════════════
# PASO 5: Importar Presupuestos desde Config_Presupuesto
# ═══════════════════════════════════════════════════════════════

def seed_presupuestos(cur, wb):
    ws = wb["Config_Presupuesto"]
    headers = [c.value for c in ws[1]]
    
    cur.execute("SELECT id, nombre FROM categorias")
    cat_map = {row[1]: row[0] for row in cur.fetchall()}
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if not d["Categoria"]:
            continue
        
        ambito = (d["Ambito"] or "").strip()
        cat = d["Categoria"].strip()
        monto = float(d["Monto"] or 0)
        cat_id = cat_map.get(cat)
        
        if not cat_id or not ambito:
            continue
        
        cur.execute("""
            INSERT INTO presupuestos (ambito, categoria_id, monto)
            VALUES (%s, %s, %s)
            ON CONFLICT (ambito, categoria_id) DO UPDATE SET
                monto = EXCLUDED.monto,
                updated_at = now()
        """, (ambito, cat_id, monto))
        count += 1
    
    print(f"✅ {count} presupuestos insertados")


# ═══════════════════════════════════════════════════════════════
# PASO 6: Importar Registros Borrados
# ═══════════════════════════════════════════════════════════════

def seed_borrados(cur, wb):
    """Inserta los registros de BaseRegistros_Borrados en gastos_borrados."""
    ws = wb["BaseRegistros_Borrados"]
    headers = [c.value for c in ws[1]]
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d["ID_Unico"] is None:
            continue
        
        cur.execute("""
            INSERT INTO gastos_borrados (
                id_unico, fecha_compra, fecha_registro,
                registrador, compra, categoria,
                monto_total, fecha_borrado
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            str(d["ID_Unico"]),
            safe_date(d["Fecha_Compra"]),
            safe_date(d["Fecha_Registro"]),
            (d["Registrador"] or ""),
            (d["Compra"] or ""),
            (d["Categoria"] or ""),
            float(d["Monto_Total"] or 0),
            safe_date(d.get("Fecha_Borrado")),
        ))
        count += 1
    
    print(f"✅ {count} registros borrados archivados")


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════

def main():
    print("🚀 Iniciando migración Finance Lab (DATOS REALES)...\n")
    
    wb = load_workbook()
    print(f"   Hojas: {wb.sheetnames}")
    
    conn = connect()
    cur = conn.cursor()
    
    # 1. Limpiar datos viejos
    clean_existing_data(cur)
    conn.commit()
    
    # 2. Cuentas
    seed_cuentas(cur, wb)
    conn.commit()
    
    # 3. Catálogos + arbolCat
    seed_catalogos(cur, wb)
    conn.commit()
    
    # 4. Gastos (desde BaseRegistros)
    seed_gastos(cur, wb)
    conn.commit()
    
    # 5. Presupuestos
    seed_presupuestos(cur, wb)
    conn.commit()
    
    # 6. Borrados
    seed_borrados(cur, wb)
    conn.commit()
    
    # ── Verificación ──
    print("\n📊 VERIFICACIÓN FINAL:")
    cur.execute("SELECT COUNT(*) FROM gastos")
    print(f"   Gastos: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM recordatorios")
    print(f"   Recordatorios: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM cuentas WHERE activo")
    print(f"   Cuentas activas: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM presupuestos WHERE activo")
    print(f"   Presupuestos: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM categoria_arbol")
    print(f"   Categoría Árbol: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM gastos_borrados")
    print(f"   Papelera: {cur.fetchone()[0]}")
    
    cur.close()
    conn.close()
    
    print("\n🎉 ¡Migración completada exitosamente!")
    print("   Datos reales de Bluecoins → PostgreSQL ✅")


if __name__ == "__main__":
    main()
